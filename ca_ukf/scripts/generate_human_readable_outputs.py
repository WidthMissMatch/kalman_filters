#!/usr/bin/env python3
"""
Generate Human-Readable VHDL Outputs
Converts Q24.24 fixed-point VHDL outputs to decimal CSV/Excel

This is the PRIMARY tool for manual inspection of VHDL UKF results.
Engineers can open the Excel file and visually scan for errors with
color-coded indicators:
  - Green: error < 0.5m (excellent)
  - Yellow: 0.5m ≤ error < 1.0m (acceptable)
  - Red: error ≥ 1.0m (investigate!)
"""

import numpy as np
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_DIR = Path(__file__).parent.parent
VHDL_OUTPUT_DIR = BASE_DIR / "results" / "vhdl_outputs" / "ghdl"
PYTHON_REF_DIR = BASE_DIR / "results" / "python_outputs" / "filterpy"
DATA_DIR = BASE_DIR / "test_data" / "real_world"
OUTPUT_DIR = BASE_DIR / "results" / "manual_inspection"

Q_SCALE = 2**24  # Q24.24 fixed-point scale

# Color coding thresholds
EXCELLENT_THRESHOLD = 0.5  # m
ACCEPTABLE_THRESHOLD = 1.0  # m

def parse_vhdl_output(vhdl_file):
    """
    Parse VHDL output text file
    
    Expected format (example):
    Cycle 0: x_pos=16777216 x_vel=0 x_acc=0 y_pos=16777216 ...
    """
    print(f"Parsing VHDL output: {vhdl_file}")
    
    cycles = []
    
    with open(vhdl_file, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or not line.startswith('Cycle'):
                continue
            
            # Parse "Cycle N: ..."
            parts = line.split(':')
            if len(parts) < 2:
                continue
            
            cycle_num = int(parts[0].replace('Cycle', '').strip())
            
            # Parse state values
            data_str = parts[1].strip()
            
            state_dict = {'cycle': cycle_num}
            
            # Extract values (format: key=value key=value ...)
            for item in data_str.split():
                if '=' in item:
                    key, val = item.split('=')
                    try:
                        # VHDL outputs Q24.24 integers
                        q24_value = int(val)
                        decimal_value = q24_value / Q_SCALE
                        state_dict[key] = decimal_value
                        state_dict[f'{key}_q24'] = q24_value
                    except ValueError:
                        continue
            
            cycles.append(state_dict)
    
    if not cycles:
        raise ValueError(f"No data parsed from {vhdl_file}")
    
    return pd.DataFrame(cycles)

def load_ground_truth(dataset_name):
    """Load ground truth trajectory"""
    csv_file = DATA_DIR / f"{dataset_name}.csv"
    if not csv_file.exists():
        raise FileNotFoundError(f"Ground truth not found: {csv_file}")
    
    return pd.read_csv(csv_file)

def load_python_reference(dataset_name):
    """Load Python UKF reference outputs"""
    pattern = f"filterpy_{dataset_name.replace('.csv', '')}*.csv"
    matches = list(PYTHON_REF_DIR.glob(pattern))
    
    if not matches:
        print(f"  Warning: No Python reference found for {dataset_name}")
        return None
    
    return pd.read_csv(matches[0])

def create_comparison_table(vhdl_df, ground_truth, python_ref=None):
    """Create side-by-side comparison table"""
    
    # Ensure cycle alignment
    min_cycles = min(len(vhdl_df), len(ground_truth))
    if python_ref is not None:
        min_cycles = min(min_cycles, len(python_ref))
    
    comparison = []
    
    for k in range(min_cycles):
        vhdl_row = vhdl_df.iloc[k]
        gt_row = ground_truth.iloc[k]
        
        row = {
            'cycle': k,
            'time': gt_row['time'],
            
            # VHDL estimates
            'vhdl_x_pos': vhdl_row.get('x_pos', np.nan),
            'vhdl_y_pos': vhdl_row.get('y_pos', np.nan),
            'vhdl_z_pos': vhdl_row.get('z_pos', np.nan),
            'vhdl_x_vel': vhdl_row.get('x_vel', np.nan),
            'vhdl_y_vel': vhdl_row.get('y_vel', np.nan),
            'vhdl_z_vel': vhdl_row.get('z_vel', np.nan),
            
            # Ground truth
            'gt_x_pos': gt_row['gt_x_pos'],
            'gt_y_pos': gt_row['gt_y_pos'],
            'gt_z_pos': gt_row['gt_z_pos'],
            'gt_x_vel': gt_row['gt_x_vel'],
            'gt_y_vel': gt_row['gt_y_vel'],
            'gt_z_vel': gt_row['gt_z_vel'],
            
            # VHDL errors
            'err_x_pos': vhdl_row.get('x_pos', np.nan) - gt_row['gt_x_pos'],
            'err_y_pos': vhdl_row.get('y_pos', np.nan) - gt_row['gt_y_pos'],
            'err_z_pos': vhdl_row.get('z_pos', np.nan) - gt_row['gt_z_pos']
        }
        
        # Position magnitude error
        row['err_pos_mag'] = np.sqrt(
            row['err_x_pos']**2 + row['err_y_pos']**2 + row['err_z_pos']**2
        )
        
        # Python reference if available
        if python_ref is not None:
            py_row = python_ref.iloc[k]
            row['py_x_pos'] = py_row['est_x_pos']
            row['py_y_pos'] = py_row['est_y_pos']
            row['py_z_pos'] = py_row['est_z_pos']
            
            # VHDL vs Python difference
            row['diff_vhdl_py_x'] = row['vhdl_x_pos'] - row['py_x_pos']
            row['diff_vhdl_py_y'] = row['vhdl_y_pos'] - row['py_y_pos']
            row['diff_vhdl_py_z'] = row['vhdl_z_pos'] - row['py_z_pos']
            row['diff_vhdl_py_mag'] = np.sqrt(
                row['diff_vhdl_py_x']**2 + row['diff_vhdl_py_y']**2 + row['diff_vhdl_py_z']**2
            )
        
        comparison.append(row)
    
    return pd.DataFrame(comparison)

def export_to_excel_with_colors(comparison_df, output_file):
    """Export to Excel with color-coded error cells"""
    print(f"Exporting to Excel with color coding: {output_file}")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VHDL Output Inspection"
    
    # Write header
    headers = list(comparison_df.columns)
    ws.append(headers)
    
    # Define color fills
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Write data with color coding
    for idx, row in comparison_df.iterrows():
        excel_row = []
        for col in headers:
            excel_row.append(row[col])
        
        ws.append(excel_row)
        
        # Color code error columns
        error_mag = row['err_pos_mag']
        excel_row_idx = idx + 2  # +1 for header, +1 for 1-indexing
        
        # Find error columns
        error_cols = [i for i, h in enumerate(headers, 1) if h.startswith('err_')]
        
        for col_idx in error_cols:
            cell = ws.cell(row=excel_row_idx, column=col_idx)
            
            if error_mag < EXCELLENT_THRESHOLD:
                cell.fill = green_fill
            elif error_mag < ACCEPTABLE_THRESHOLD:
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save
    wb.save(output_file)
    print(f"  ✓ Excel file created with {len(comparison_df)} cycles")

def generate_summary_stats(comparison_df):
    """Generate summary statistics for the comparison"""
    stats = {
        'total_cycles': len(comparison_df),
        'pos_rmse': np.sqrt(np.mean(comparison_df['err_pos_mag']**2)),
        'pos_max_error': comparison_df['err_pos_mag'].max(),
        'pos_mean_error': comparison_df['err_pos_mag'].mean(),
        'excellent_cycles': (comparison_df['err_pos_mag'] < EXCELLENT_THRESHOLD).sum(),
        'acceptable_cycles': ((comparison_df['err_pos_mag'] >= EXCELLENT_THRESHOLD) & 
                              (comparison_df['err_pos_mag'] < ACCEPTABLE_THRESHOLD)).sum(),
        'problematic_cycles': (comparison_df['err_pos_mag'] >= ACCEPTABLE_THRESHOLD).sum()
    }
    
    stats['excellent_pct'] = stats['excellent_cycles'] / stats['total_cycles'] * 100
    stats['acceptable_pct'] = stats['acceptable_cycles'] / stats['total_cycles'] * 100
    stats['problematic_pct'] = stats['problematic_cycles'] / stats['total_cycles'] * 100
    
    return stats

def print_summary(stats):
    """Print summary statistics"""
    print(f"\n{'='*80}")
    print("VHDL OUTPUT SUMMARY")
    print('='*80)
    print(f"Total cycles: {stats['total_cycles']}")
    print(f"\nPosition Error:")
    print(f"  RMSE:       {stats['pos_rmse']:.4f} m")
    print(f"  Mean:       {stats['pos_mean_error']:.4f} m")
    print(f"  Max:        {stats['pos_max_error']:.4f} m")
    print(f"\nError Distribution:")
    print(f"  Excellent (< {EXCELLENT_THRESHOLD}m):  {stats['excellent_cycles']:4d} ({stats['excellent_pct']:5.1f}%)")
    print(f"  Acceptable ({EXCELLENT_THRESHOLD}-{ACCEPTABLE_THRESHOLD}m): {stats['acceptable_cycles']:4d} ({stats['acceptable_pct']:5.1f}%)")
    print(f"  Problematic (≥ {ACCEPTABLE_THRESHOLD}m): {stats['problematic_cycles']:4d} ({stats['problematic_pct']:5.1f}%)")
    
    if stats['problematic_pct'] > 5:
        print(f"\n⚠ WARNING: {stats['problematic_pct']:.1f}% of cycles have errors ≥ {ACCEPTABLE_THRESHOLD}m")
        print("  Manual inspection recommended")
    elif stats['excellent_pct'] > 95:
        print(f"\n✓ EXCELLENT: {stats['excellent_pct']:.1f}% of cycles have errors < {EXCELLENT_THRESHOLD}m")
    else:
        print(f"\n✓ GOOD: Most cycles within acceptable bounds")

def process_vhdl_output(vhdl_file, dataset_name):
    """Process one VHDL output file"""
    print(f"\n{'='*80}")
    print(f"Processing: {vhdl_file.name}")
    print('='*80)
    
    # Parse VHDL output
    vhdl_df = parse_vhdl_output(vhdl_file)
    print(f"Parsed {len(vhdl_df)} cycles from VHDL output")
    
    # Load ground truth
    ground_truth = load_ground_truth(dataset_name)
    
    # Load Python reference (optional)
    python_ref = load_python_reference(dataset_name)
    
    # Create comparison
    comparison_df = create_comparison_table(vhdl_df, ground_truth, python_ref)
    
    # Generate summary
    stats = generate_summary_stats(comparison_df)
    print_summary(stats)
    
    # Export CSV
    csv_file = OUTPUT_DIR / f"vhdl_readable_{dataset_name}.csv"
    comparison_df.to_csv(csv_file, index=False, float_format='%.6f')
    print(f"\n✓ CSV saved: {csv_file.name}")
    
    # Export Excel with colors
    excel_file = OUTPUT_DIR / f"vhdl_readable_{dataset_name}.xlsx"
    export_to_excel_with_colors(comparison_df, excel_file)
    print(f"✓ Excel saved: {excel_file.name}")
    
    return comparison_df, stats

def main():
    print("="*80)
    print("VHDL OUTPUT → HUMAN-READABLE CONVERTER")
    print("="*80)
    print("Converts Q24.24 fixed-point to decimal CSV/Excel")
    print("Color coding:")
    print(f"  🟢 Green:  error < {EXCELLENT_THRESHOLD}m (excellent)")
    print(f"  🟡 Yellow: {EXCELLENT_THRESHOLD}m ≤ error < {ACCEPTABLE_THRESHOLD}m (acceptable)")
    print(f"  🔴 Red:    error ≥ {ACCEPTABLE_THRESHOLD}m (investigate)")
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Find VHDL output files
    vhdl_files = list(VHDL_OUTPUT_DIR.glob("*.txt"))
    
    if not vhdl_files:
        print(f"\n✗ No VHDL outputs found in {VHDL_OUTPUT_DIR}")
        print("Run GHDL simulations first")
        return
    
    print(f"\nFound {len(vhdl_files)} VHDL output files")
    
    all_stats = []
    
    for vhdl_file in vhdl_files:
        # Try to infer dataset name from filename
        # Expected: vhdl_drone_euroc_mh01.txt → drone_euroc_mh01
        dataset_name = vhdl_file.stem.replace('vhdl_', '')
        
        try:
            comparison_df, stats = process_vhdl_output(vhdl_file, dataset_name)
            stats['dataset'] = dataset_name
            stats['vhdl_file'] = vhdl_file.name
            all_stats.append(stats)
        except Exception as e:
            print(f"\n✗ Error processing {vhdl_file.name}: {e}")
            import traceback
            traceback.print_exc()
    
    # Overall summary
    if all_stats:
        print(f"\n{'='*80}")
        print("OVERALL SUMMARY")
        print('='*80)
        
        summary_df = pd.DataFrame(all_stats)
        print(f"\nProcessed {len(summary_df)} VHDL outputs")
        print(f"Average RMSE: {summary_df['pos_rmse'].mean():.4f} m")
        print(f"Average excellent rate: {summary_df['excellent_pct'].mean():.1f}%")
        
        csv_file = OUTPUT_DIR / "vhdl_summary.csv"
        summary_df.to_csv(csv_file, index=False, float_format='%.4f')
        print(f"\n✓ Summary saved: {csv_file}")
    
    print(f"\n{'='*80}")
    print("CONVERSION COMPLETE")
    print('='*80)
    print(f"Output directory: {OUTPUT_DIR}")
    print("\nNext steps:")
    print("  1. Open Excel files for visual inspection")
    print("  2. Run interactive_output_viewer.py for detailed analysis")
    print("  3. Run manual_comparison_tool.py for specific cycle investigation")

if __name__ == "__main__":
    main()
